"""Excel-Export: ein Tabellenblatt pro Tag (Tag oder Datumsbereich)."""
from __future__ import annotations

import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import storage
from ..service import analysis

_HEADER_FILL = PatternFill("solid", fgColor="2E5E3A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MEAL_FILL = PatternFill("solid", fgColor="D7E8DC")
_TOTAL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COLS = ["Mahlzeit", "Menge", "Produkt", "kcal", "Fett (g)", "KH (g)", "Eiweiß (g)"]


def _safe_sheet_title(title: str) -> str:
    for ch in r"[]:*?/\\":
        title = title.replace(ch, "-")
    return title[:31]


def _write_day_sheet(ws, day_view: dict, profile: dict) -> None:
    row = 1
    ws.cell(row=row, column=1, value=f"FDDB Tagesauswertung - {day_view['date']} "
            f"({day_view['weekday']})").font = Font(bold=True, size=13)
    row += 1
    if profile.get("name"):
        ws.cell(row=row, column=1, value=f"Profil: {profile['name']}"
                + (f", geb. {profile['birthdate']}" if profile.get('birthdate') else ""))
        row += 1
    note = day_view.get("note")
    if note:
        ws.cell(row=row, column=1, value=f"Besonderheiten: {note}")
        row += 1
    row += 1

    header_row = row
    for c, name in enumerate(_COLS, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
    row += 1

    for meal in day_view["meals"]:
        for p in meal["products"]:
            values = [meal["name"], p["amount"], p["name"], p["kcal"],
                      p["fat"], p["carbs"], p["protein"]]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = _BORDER
                if c == 1:
                    cell.fill = _MEAL_FILL
            row += 1
        # Mahlzeit-Zwischensumme
        subtotal = [meal["name"] + " gesamt", "", "", meal["kcal"],
                    meal["fat"], meal["carbs"], meal["protein"]]
        for c, v in enumerate(subtotal, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = _TOTAL_FONT
            cell.fill = _MEAL_FILL
            cell.border = _BORDER
        row += 1

    t = day_view["totals"]
    total = ["Tagessumme", "", "", t.get("kcal"), t.get("fat"),
             t.get("carbs"), t.get("protein")]
    for c, v in enumerate(total, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(bold=True, size=12)
        cell.border = _BORDER
    row += 1
    extra = []
    if t.get("sugar") is not None:
        extra.append(f"davon Zucker: {t['sugar']} g")
    if t.get("fiber") is not None:
        extra.append(f"Ballaststoffe: {t['fiber']} g")
    if extra:
        ws.cell(row=row, column=1, value="   ".join(extra))

    widths = [16, 12, 40, 10, 10, 10, 11]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def export_days(date_isos: List[str]) -> bytes:
    """Erzeugt eine Excel-Datei mit je einem Blatt pro Tag."""
    profile = storage.load_settings()
    wb = Workbook()
    wb.remove(wb.active)
    if not date_isos:
        ws = wb.create_sheet("Leer")
        ws["A1"] = "Keine Daten fuer den gewaehlten Zeitraum."
    for iso in date_isos:
        day_view = analysis.get_day(iso)
        if day_view is None:
            continue
        ws = wb.create_sheet(_safe_sheet_title(iso))
        _write_day_sheet(ws, day_view, profile)
    if not wb.sheetnames:
        ws = wb.create_sheet("Leer")
        ws["A1"] = "Keine Daten."
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_single_day(date_iso: str) -> bytes:
    return export_days([date_iso])


def export_range(from_iso: str, to_iso: str) -> bytes:
    rng = analysis.get_range(from_iso, to_iso)
    return export_days([d["date"] for d in rng["days"]])
